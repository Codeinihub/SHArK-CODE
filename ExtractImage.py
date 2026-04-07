import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from PIL import Image

# --- CONFIGURATION (Default Ratios for 560x420) ---
# These are percentages (0.0 to 1.0) of the width and height
# Use these if you DON'T want to click every time
DEFAULT_RATIOS = {
    "cb_x": 0.84,      # Colorbar X center
    "cb_y_top": 0.08,   # 5.0 value
    "cb_y_bot": 0.90,   # 0.0 value
}

def get_manual_corners(img):
    """Pops up the image for you to click the 4 grid corners."""
    plt.imshow(img)
    plt.title("Click 4 corners: 1.Back-Left, 2.Back-Right, 3.Front-Left, 4.Front-Right")
    # points will be [(x1, y1), (x2, y2), (x3, y3), (x4, y4)]
    points = plt.ginput(4, timeout=0) 
    plt.close()
    
    # Convert to ratios so it works if the image size changes later
    w, h = img.size
    ratios = {
        (0, 0): np.array([points[0][0]/w, points[0][1]/h]),
        (0, 7): np.array([points[1][0]/w, points[1][1]/h]),
        (7, 0): np.array([points[2][0]/w, points[2][1]/h]),
        (7, 7): np.array([points[3][0]/w, points[3][1]/h]),
    }
    return ratios

def extract_from_image_interactive(image_path):
    img = Image.open(image_path).convert('RGB')
    w, h = img.size
    arr = np.array(img)

    # 1. GET CORNER RATIOS (Manual or Default)
    # Change to False if you want to use hardcoded defaults
    INTERACTIVE_MODE = True 
    if INTERACTIVE_MODE:
        corner_ratios = get_manual_corners(img)
    else:
        # Fallback to defaults (you'd define these like the interactive ones)
        print("Using default corner ratios...")

    # 2. BUILD COLORBAR LOOKUP (Using ratios)
    cb_x = int(DEFAULT_RATIOS["cb_x"] * w)
    y1 = int(DEFAULT_RATIOS["cb_y_top"] * h)
    y2 = int(DEFAULT_RATIOS["cb_y_bot"] * h)
    
    colorbar_slice = arr[y1:y2, cb_x, :]
    # The bar goes 5.0 (top) to 0.0 (bottom)
    colorbar_values = np.linspace(5.0, 0.0, len(colorbar_slice))

    # 3. INTERPOLATE AND SAMPLE
    grid = np.zeros((8, 8))
    
    for r in range(8):
        for c in range(8):
            # Bilinear interpolation using the ratios
            t, s = r/7.0, c/7.0
            p00, p01 = corner_ratios[(0,0)], corner_ratios[(0,7)]
            p10, p11 = corner_ratios[(7,0)], corner_ratios[(7,7)]
            
            rel_pos = (1-t)*(1-s)*p00 + (1-t)*s*p01 + t*(1-s)*p10 + t*s*p11
            px, py = int(rel_pos[0] * w), int(rel_pos[1] * h)

            # Sample 3x3 patch to get a clean color (avoids grid lines)
            patch = arr[py-1:py+2, px-1:px+2, :]
            avg_color = np.mean(patch, axis=(0,1))

            # Find closest value in colorbar
            diffs = np.linalg.norm(colorbar_slice - avg_color, axis=1)
            grid[r, c] = colorbar_values[np.argmin(diffs)]

    return grid
def append_grid_to_excel(grid, catalyst_name, excel_path):
    """
    Appends the 8x8 grid to the Excel file in the exact format 
    your FeatExtract.py script expects (Space + Name + 8x8 Grid).
    """
    # 1. Load the existing workbook
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # 2. Find the last row with data
    last_row = ws.max_row
    
    # 3. Leave one blank row (Row N+1)
    start_row = last_row + 2 
    
    # 4. Write the Catalyst Name (Row N+2, Column B to match your image)
    ws.cell(row=start_row, column=2, value=catalyst_name)
    ws.cell(row=start_row, column=3, value="Extracted from Image") # Optional note
    
    # 5. Write the 8x8 Grid (Starting from Row N+3)
    for r in range(8):
        for c in range(8):
            # Write to Columns A-H (1-8)
            ws.cell(row=start_row + 1 + r, column=c + 1, value=grid[r, c])
            
    # 6. Save the file
    wb.save(excel_path)
    print(f" Successfully 'printed' {catalyst_name} into {excel_path}")

# Example usage:
grid = extract_from_image_interactive(r"C:\Users\User\OneDrive\Dokumenty\SHArK Images\YKMn.jpg")
# save_to_master_dataset(grid, "TaZnFe")
append_grid_to_excel(grid, "YKMn", r"C:\Users\User\OneDrive\Dokumenty\Copy of shark data.xlsx")

print(grid)
